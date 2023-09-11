"""Self_Audit model class."""
from __future__ import annotations

import os
import time

import openpyxl as openpyxl
import sqlalchemy as db
from sqlalchemy.orm import sessionmaker

from model.entity.policyrequirement_db.audit_categories import AuditCategory
from model.entity.policyrequirement_db.policy_compliances import PolicyCompliance
from model.entity.policyrequirement_db.policy_requirement_license_types import PolicyRequirementLicenseType
from model.entity.policyrequirement_db.policy_requirement_permits import PolicyRequirementPermit
from model.entity.policyrequirement_db.policy_requirement_verticals import PolicyRequirementVertical
from model.entity.policyrequirement_db.policy_requirements import PolicyRequirement
from model.entity.policyrequirement_db.policy_vertical_techniques import PolicyVerticalTechnique
from model.entity.policyrequirement_db.regulation_locations import RegulationLocation
from model.repository.db_model import DBModel
from model.repository.license_model import LicenseModel
from model.repository.organization_db_model import OrganizationDBModel
from service.sheet_column_service import SheetColumnService
from util.app_logger import AppLogger


class SelfAuditModel(DBModel):

    """SelfAudit class."""

    def __init__(self) -> None:
        """To initialize the Constructor method."""
        super().__init__()
        self.logger = AppLogger.get_instance()
        self.connection = self.connect('PolicyRequirementDB')
        self.session_maker = sessionmaker(bind=self.engine)
        self.session = self.session_maker()
        self.sheet_service = SheetColumnService()
        self.license_model = LicenseModel()
        self.organization_model = OrganizationDBModel()
        self.session.flush()

    def export_self_audit(self, city: str, state: str, county: str, client_name: str, organization_id: int) -> str:
        """To export self audit data in excel file."""
        policy_requirements = self.get_state_policy_requirements(
            city, state, county, organization_id)
        self_audit_data = []
        for index, policy_requirement in enumerate(policy_requirements):
            export_data = []
            policy_compliance = self.get_policy_compliance(
                policy_requirement.policy_requirement_id)
            policy_permits = self.get_policy_requirement_permits(
                policy_requirement.policy_requirement_id)
            policy_verticals = self.get_policy_requirement_verticals(
                policy_requirement.policy_requirement_id)
            policy_license_types = self.get_policy_requirement_license_types(
                policy_requirement.policy_requirement_id)
            export_data.append(policy_requirement.policy_requirement_id)
            if policy_requirement.organization_id is not None and policy_requirement.organization_id > 0:
                org_name = self.organization_model.get_active_organization_name(
                    policy_requirement.organization_id)
                export_data.append(org_name)
            else:
                export_data.append(client_name)

            export_data.append(policy_requirement.city)
            export_data.append(policy_requirement.county)
            export_data.append(policy_requirement.state)
            export_data.append(policy_requirement.code)
            export_data.append(policy_requirement.chapter)
            export_data.append(policy_requirement.section)
            export_data.append(policy_requirement.regulation)
            export_data.append(policy_requirement.category)
            export_data.append(policy_requirement.question_order)
            if policy_compliance:
                for row in policy_compliance:
                    if (row.level == '1'):
                        export_data.append(row.question)
                        if row.non_compliant.lower() == 'yes':
                            export_data.append('No')
                        else:
                            export_data.append('Yes')
                        export_data.append(row.action)
                        if len(policy_compliance) == 1:
                            export_data.append('')
                            export_data.append('')
                            export_data.append('')
                            export_data.append('')
                    else:
                        export_data.append(row.trigger_response)
                        export_data.append(row.question)
                        if row.non_compliant.lower() == 'yes':
                            export_data.append('No')
                        else:
                            export_data.append('Yes')
                        export_data.append(row.action)

                export_data.append(policy_requirement.user_facing_note)
                if policy_license_types:
                    export_data.append(
                        self.license_model.get_license_type_names(policy_license_types))
                else:
                    export_data.append('')
                if policy_verticals:
                    export_data.append(
                        self.license_model.get_license_vertical_names(policy_verticals))
                else:
                    export_data.append('')
                if policy_requirement.recreational == 1 and policy_requirement.medicinal == 1:
                    export_data.append('Both')
                elif policy_requirement.recreational == 1:
                    export_data.append('Recreational')
                else:
                    export_data.append('Medicinal')
                if policy_permits:
                    export_data.append(
                        self.license_model.get_license_permit_names(policy_permits))
                else:
                    export_data.append('')
                if policy_verticals:
                    policy_requirement_vertical_id = self.get_policy_requirement_verticals_technics(
                        policy_requirement.policy_requirement_id)
                    policy_techniques = self.get_policy_vertical_techniques(
                        policy_requirement_vertical_id)
                    if policy_techniques:
                        export_data.append(
                            self.license_model.get_license_vertical_technique_names(policy_techniques))
                    else:
                        export_data.append('')
                else:
                    export_data.append('')
                export_data.append('')

            self_audit_data.append(export_data)
        if len(policy_requirements) > 0:
            file_name = self.create_file_name(city, state, county, client_name)
            return self.self_audit_file_Save(self_audit_data, file_name)

    def create_file_name(self, city: str, state: str, county: str, client_name: str) -> str:
        """To generate export file name."""
        if client_name:
            file_name = 'SAQs_'+client_name+'_'+state
        else:
            file_name = 'SAQs_'+state
        if city:
            file_name += '_'+city
        elif county:
            file_name += '_'+county
        timestr = '_' + time.strftime('%m%d%Y')
        file_name += timestr + '.xlsx'
        return file_name

    def set_export_file_path(self, dest_path: str, file_name: str) -> str:
        """Set the export file path for new file."""
        if not os.path.exists(dest_path):
            os.makedirs(dest_path)

        return dest_path+file_name

    def self_audit_file_Save(self, selfAudit_Data: list, file_name: str) -> str:
        """To copy self audit data in excel file."""
        # opening the source excel file
        source_path = r'/app/Templates/SAQs_Client_Name_State Name_CityCounty Name_MMDDYYYY.xlsx'
        dest_path = r'/app/Export_SelfAudit/'
        export_path = self.set_export_file_path(dest_path, file_name)
        wb1 = openpyxl.load_workbook(source_path)
        wb1.save(export_path)
        # opening the destination excel file
        wb2 = openpyxl.load_workbook(export_path)
        ws2 = wb2.active
        # calculate total number of rows and
        # columns in source excel file
        mr = len(selfAudit_Data)
        mc = len(self.sheet_service.selfaudit_column())

        # copying the cell values from source
        # excel file to destination excel file
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = selfAudit_Data[i-1][j-1]
                # writing the read value to destination excel file
                ws2.cell(row=i+1, column=j).value = c
        # saving the destination excel file
        wb2.save(str(export_path))
        return export_path

    def get_state_policy_requirements(self, city: str, state: str, county: str, organization_id: int) -> list:
        """Get policy requirements based on state,city and county."""
        result = self.session.query(
            PolicyRequirement.medicinal,
            PolicyRequirement.recreational,
            PolicyRequirement.policy_requirement_id,
            PolicyRequirement.section,
            PolicyRequirement.code,
            PolicyRequirement.chapter,
            PolicyRequirement.section,
            PolicyRequirement.regulation,
            PolicyRequirement.user_facing_note,
            PolicyRequirement.question_order,
            PolicyRequirement.audit_category_id,
            RegulationLocation.city,
            RegulationLocation.county,
            RegulationLocation.state,
            AuditCategory.category,
            AuditCategory.organization_id,
        ).join(RegulationLocation,
               PolicyRequirement.regulation_location_id == RegulationLocation.regulation_location_id
               ).join(AuditCategory,
                      AuditCategory.audit_category_id == PolicyRequirement.audit_category_id).where(
            RegulationLocation.city == city,
            RegulationLocation.state == state,
            RegulationLocation.county == county,
            PolicyRequirement.is_active == 1,
        )
        if organization_id > 0:
            result = result.filter(
                AuditCategory.organization_id == organization_id).all()
        else:
            result = result.all()
        return result

    def get_policy_compliance(self, policy_req_id: int) -> list:
        """Get policy compliance based on policy_req_id."""
        query = db.select([
            PolicyCompliance.policy_compliance_id,
            PolicyCompliance.question,
            PolicyCompliance.non_compliant,
            PolicyCompliance.action,
            PolicyCompliance.trigger_response,
            PolicyCompliance.level

        ]).where(
            PolicyCompliance.policy_requirement_id == policy_req_id,
        )
        return self.connection.execute(query).fetchall()

    def get_policy_vertical_techniques(self, policy_req_vertical_id: list) -> list:
        """Get policy vertical technique based on policy_req_vertical_id."""
        query = db.select([
            PolicyVerticalTechnique.license_type_vertical_technique_id
        ]).distinct().where(
            PolicyVerticalTechnique.policy_requirement_vertical_id.in_(policy_req_vertical_id))
        resp = self.connection.execute(query).fetchall()
        return [id[0] for id in resp]

    def get_policy_requirement_verticals(self, policy_req_id: int) -> list:
        """Get policy verticals based on policy_req_id."""
        query = db.select([
            PolicyRequirementVertical.license_type_vertical_id,
        ]).distinct().where(
            PolicyRequirementVertical.policy_requirement_id == policy_req_id)
        resp = self.connection.execute(query).fetchall()
        return [id[0] for id in resp]

    def get_policy_requirement_verticals_technics(self, policy_req_id: int) -> list:
        """Get policy verticals technics based on policy_req_id."""
        query = db.select([
            PolicyRequirementVertical.policy_requirement_vertical_id
        ]).distinct().where(
            PolicyRequirementVertical.policy_requirement_id == policy_req_id)
        resp = self.connection.execute(query).fetchall()
        return [id[0] for id in resp]

    def get_policy_requirement_permits(self, policy_req_id: int) -> list:
        """Get policy permits based on policy_req_id."""
        query = db.select([
            PolicyRequirementPermit.license_type_permit_id,
        ]).distinct().where(
            PolicyRequirementPermit.policy_requirement_id == policy_req_id)
        resp = self.connection.execute(query).fetchall()
        return [id[0] for id in resp]

    def get_policy_requirement_license_types(self, policy_req_id: int) -> list:
        """Get policy license types based on policy_req_id."""
        query = db.select([
            PolicyRequirementLicenseType.license_type_id,
        ]).distinct().where(
            PolicyRequirementLicenseType.policy_requirement_id == policy_req_id)
        resp = self.connection.execute(query).fetchall()
        return [id[0] for id in resp]

    def get_audit_categories(self, audit_category_id: int) -> str:
        """Get audit category based on audit_category_id."""
        query = db.select([
            AuditCategory.category,
        ]).where(
            AuditCategory.audit_category_id == audit_category_id)
        return self.connection.execute(query).scalar()
